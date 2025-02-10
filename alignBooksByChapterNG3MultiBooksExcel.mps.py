import sys, os
import regex as re
from datetime import datetime, timedelta
from pathlib import Path, PurePath
from math import ceil
from random import seed as seed
import numpy as np
import sqlite3
from tqdm import tqdm 
import torch
from sentence_splitter import SentenceSplitter, split_text_into_sentences
#from nltk import word_tokenize
import unicodedata
import pysbd
import opencc

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows


from dp_utils import make_alignment_types, read_alignments, \
    read_in_embeddings, make_doc_embedding, vecalign, yield_overlaps

from score import score_multiple, log_final_scores

from sentence_transformers import SentenceTransformer, models, util

s2tw = opencc.OpenCC('s2tw.json')


start_time = datetime.now()
if os.name == 'nt':
    d = 1
elif os.name == 'posix':
    if torch.cuda.is_available():
        d = 0
    elif torch.mps.is_available():
        d = 1
dev = ['cuda', 'mps', 'cpu'][d]
#Model we want to use for bitext mining. LaBSE achieves state-of-the-art performance

m = 0
model_name = ['ibm-granite/granite-embedding-278m-multilingual', 'LaBSE', 'paraphrase-multilingual-MiniLM-L12-v2'][m]
#model_name = "/Users/rubentsui/.cache/lm-studio/models/lmstudio-community/granite-embedding-278m-multilingual-GGUF/granite-embedding-278m-multilingual-Q8_0.gguf"  # Your GGUF file


if 'model' not in globals():
    print(f"Now running bitext mining with transformer model [{model_name}] on device [{dev}]...", flush=True)
    model = SentenceTransformer(model_name, device=dev)
    print(f"Finished loading model: {model_name}.", flush=True)
else:
    print(f"Model [{model_name}] already loaded", flush=True)

end_time = datetime.now() - start_time
print(f"Model-loading time: {end_time.seconds} secs", flush=True)

#%%

def encodeVectors(ss, model, dev):
    '''
    Input:
        ss: lit of strings 
    Output:
        pytorch tensors
    '''
    vecs = model.encode(ss, show_progress_bar=False, convert_to_numpy=False, normalize_embeddings=True, device=dev)
    return vecs

def print_alignments(alignments, scores=None, file=sys.stdout):
    if scores is not None:
        for (x, y), s in zip(alignments, scores):
            print('%s:%s:%.6f' % (x, y, s), file=file)
    else:
        for x, y in alignments:
            print('%s:%s' % (x, y), file=file)


def file_open(filepath):
    #Function to allowing opening files based on file extension
    if filepath.endswith('.gz'):
        return gzip.open(filepath, 'rt', encoding='utf8')
    elif filepath.endswith('.bz2'):
        return bz2.open(filepath, 'rt', encoding='utf8')
    elif filepath.endswith('.xz'):
        return lzma.open(filepath, 'rt', encoding='utf8')
    else:
        return open(filepath, 'r', encoding='utf8')

def getLines(fin):
    '''
    Retrive lines from a file or (later) sqlite3 database
    '''
    lines = file_open(fin).readlines()
    return [s.strip() for s in lines if s.strip() != '']

def getSentIndex(lines):
    """
    dictionary look-up:
        keys = sentence or overlapped sentences
        value = index
    """
    sent2line = dict()
    for ii, line in enumerate(lines):
        if line.strip() in sent2line:
            raise Exception('got multiple embeddings for the same line')
        sent2line[line.strip()] = ii
    return sent2line

def getOverlaps(lines, num_overlaps):
    output = set()
    for out_line in yield_overlaps(lines, num_overlaps):
        output.add(out_line)

    # for reproducibility
    output = list(output)
    output.sort()
    return output

def normalizeText(text):
    text = text.replace("\xad", '')  # remove Unicode soft hyphen
    return unicodedata.normalize("NFKC", text) # remove Unicode , among others

# Sentence tokenizer

# regex to identify Chinese sentence boundaries
#regex_zh_sent_delim = re.compile(r"([。！？；][」』”〕》〗】)）\]]?)")
#regex_zh_sent_delim = re.compile(r"([。？；][」』”〕》〗】)）\]]?)")
#regex_zh_sent_delim = re.compile(r'(?P<quotation_mark>([。？！…]{1,2})[」』〕》〗】\])”’"\'）])')
#regex_zh_sent_delim = re.compile(r"[。！？]")
regex_zh_sent_delim = re.compile(r"([。？！…][」』”’\'\"〕》〗】)）\]]{0,3})")

def normalizeTextZh(text):
    text = text.replace("\xad", '')  # remove Unicode 
    #text = text.replace("!", "！").replace(";", "；")
    return unicodedata.normalize("NFKD", text) # remove Unicode , among others

def sentencizeZh(s):
    '''
    turn long string s into a list of sentences
    '''
    s = normalizeTextZh(s)
    s = s.replace(',','，').replace(';','；').replace("!", "！").replace(":", "：").replace("?", "？")
    ss = regex_zh_sent_delim.sub(r"\1\n", s).split("\n")
    return [s.strip() for s in ss if s.strip() != '']


def sentencize(s, lang='en'):
    if lang in ['zh', 'ja']:
        return sentencizeZh(s)
    else: # lang in ['en', 'es', 'fr', 'de', 'it', etc. ]
        splitter = SentenceSplitter(language=lang)
        sentseg = pysbd.Segmenter(language=lang, clean=False)
        s = normalizeText(s)
        ss = splitter.split(text=s)
        #ss = sentseg.segment(s)
        return [s.strip() for s in ss if s.strip() != '']

def convertChinesePunctuations(txt):
    '''
    Convert “”‘’ to, respeectively 「」『』 
    '''
    punctHans2Hant = {'“”‘’': '「」『』'}
    for k in punctHans2Hant:
        v = punctHans2Hant[k]
        for ps, pt in zip(k, v):
            txt = txt.replace(ps, pt)
    return txt
    

def align(sS, sT, alignment_max_size=4):
     
    # make runs consistent
    seed(42)
    np.random.seed(42)

    # source
    overlapsS = getOverlaps(sS, alignment_max_size)  # create "overlapped" sentences
    s2idxS = getSentIndex(overlapsS)                 # create "sentence-to-index" lookup table
    embedS = encodeVectors(overlapsS, model, dev)    # encode a list of sentences 
    src_line_embeddings = torch.vstack(embedS).cpu().numpy()   # turns a list of sentences into a tensor object
    # target
    overlapsT = getOverlaps(sT, alignment_max_size)
    s2idxT = getSentIndex(overlapsT)
    embedT = encodeVectors(overlapsT, model, dev)
    overlapsS = getOverlaps(sS, alignment_max_size)
    tgt_line_embeddings = torch.vstack(embedT).cpu().numpy()
    
    #print(f"src_line_embeddings has shape: [{src_line_embeddings.shape}]")
    #print(f"tgt_line_embeddings has shape: [{tgt_line_embeddings.shape}]")
    #sys.exit(0)

    width_over2 = ceil(alignment_max_size / 2.0) + 5

    test_alignments = []
    stack_list = []
    
    #src_lines = open(finS, 'rt', encoding="utf-8").readlines()
    vecs0 = make_doc_embedding(s2idxS, src_line_embeddings, sS, alignment_max_size)

    #tgt_lines = open(finT, 'rt', encoding="utf-8").readlines()
    vecs1 = make_doc_embedding(s2idxT, tgt_line_embeddings, sT, alignment_max_size)

    final_alignment_types = make_alignment_types(alignment_max_size)

    stack = vecalign(vecs0=vecs0,
                     vecs1=vecs1,
                     final_alignment_types=final_alignment_types,
                     del_percentile_frac=0.2,
                     width_over2=width_over2,
                     max_size_full_dp=300,
                     costs_sample_size=20000,
                     num_samps_for_norm=100)

    # write final alignments to fk\ile
    #print_alignments(stack[0]['final_alignments'], stack[0]['alignment_scores'])
    #test_alignments.append(stack[0]['final_alignments'])
    #stack_list.append(stack)
    
    alignments = stack[0]['final_alignments']
    scores     = stack[0]['alignment_scores']

    aligned_sentences = []
    if scores is not None:
        for (idxS, idxT), score in zip(alignments, scores):
            sbS  = [] # sentence block - source
            for i in idxS:
                sbS.append(sS[i])            
            sbT  = [] # sentence block - target
            for i in idxT:
                sbT.append(sT[i])
            
            #aligned_sentences.append(f"{score:.5f}\t{idxS}\t{' '.join(sbS)}\t{idxT}\t{' '.join(sbT)}")            
            aligned_sentences.append([score, idxS, ' '.join(sbS), idxT, ' '.join(sbT)])      
    return aligned_sentences
#%%

def createExcel(fin):
    
    """ fin = plain text aligned text
    """
    
    col_widths = {'zh': 50, 'en': 65}
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Select the active sheet
    ws = wb.active
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = col_widths[langS]
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = col_widths[langT]
    
    data = open(fin, 'r', encoding='utf-8').readlines()

    df = pd.DataFrame([x.split('\t') for x in data], columns=['cosdist', 'cols_s', langS, 'cols_t',  langT])

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    # Set cell alignment
    alignment = Alignment(horizontal='general',
                          vertical='top',
                          wrap_text=True)
    cnt = len(data)
    for row in ws[f'A1:F{cnt+10}']:
        for cell in row:
            cell.alignment = alignment

    # Save the workbook
    base = Path(fin).stem
    fon_xlsx = Path(fin).parent / f'{base}.xlsx'
    wb.save(fon_xlsx)

#%%

if __name__ == '__main__':
    
    alignment_max_size = 9
    print(f"alignment_max_size = {alignment_max_size}")
    
    ###########################################################
    # Step 1 Use chapter separator?
    # Step 2 Convert to Traditional Chinese?
    ###########################################################
    USE_REGEX_CHAPTER_SEPARATOR = True #   False
    CONVERT_ZHS_TO_ZHT = False # True # False # True

    ###########################################################
    # Step 3 Choose language pair (translation direction)
    ###########################################################
    langS = ['zh', 'en', 'fr', 'zh', 'en', 'es', 'zh', 'ja', 'es'][1]
    langT = ['zh', 'en', 'fr', 'it', 'zhs'][0]

    out_langS, out_langT = langS, langT

    ###########################################################
    # Step 5 Choose input file folder 
    ###########################################################
    base_folder = ["HP", "./Joyce", "./GPTI", "./Chinese_French", "./Chinese_English", "./English_Chinese"][-1]

    in_folder =  f"./{base_folder}"
    out_folder = f"./{base_folder}_aligned" 

    filesS = sorted(Path(in_folder).glob(f"*.{langS}.txt"))
    filesT = sorted(Path(in_folder).glob(f"*.{langT}.txt"))

    ######################################################################
    # regex for dividing text into chunks (chapter, book, section, etc.)
    DregexS = {'': r"",
               'HP1':   r"— CHAPTER .+?",
               'HP2':   r"— CHAPTER .+?",
               'HP3':   r"- CHAPTER .+?",
               'HP4':   r"— CHAPTER .+?",
               'HP5':   r"— CHAPTER .+?",
               'HP6':   r"— CHAPTER .+?",
               'HP7':   r"— CHAPTER .+?",
               'Dubliners': r"",
               'Asimov_Foundation1': r"\d{1,2}",
               'Asimov_Foundation2': r"\d{1,2}\. +.*?",
               'Asimov_Foundation3': r"\d{1,2} +.*?",
               'Asimov_FoundationPrelude': r"\d{1,2}",
               'Asimov_FoundationEdge': r"\d{1,2}",
               'Asimov_FoundationEarth': r"Chapter \d{1,3}: .*?",
               'Capitalism': r"\d{1,2}",
               'SongIceFire1': r"[A-Z]{3,}",
               'SongIceFire2': r"[A-Z]{3,}",
               'SongIceFire3': r"[A-Z ’]{3,}",
               'SongIceFire4': r"[A-Z ’]{3,}",
               'Corton_LondonFog': r"CHAPTER .*?",
               'Taipei': r"第.{1,3}章",
               'Patten_EastWest': r"\d+\. .+?",
               'SC3a': r"Chapter \d{1,3}",
               'StoryStone': r"第.{1,4}回",
               'Vogel_Deng': r"\d{1,2}",
               'Vogel_Deng_Notes': r"Chapter \d+\..+?",
               'OrphanAsia': r"【.+?】",
               'Westad_RestlessEmpire': r"CHAPTER \d{1,3} .+",
               'Tolstoy_WarPeace': r"[IVXLCM]{1,8}",
               'Yanagihara_PeopleTrees_TextPlusNotes': r"PART\s+[IVX]+[\.:]\s+.+",
               'DMS1': r"\d{1,3}",
               'DMS2': r"\d{1,3}",
               'DMS3': r"\d{1,3}",
               'Dimbleby_LastGovernor': r"\d{1,2}",
               'Crystal': r"第.部",
               'GoodEarth': r"[\dI]{1,2}",
               'Oriental': r"\d\..+",
               'CondorHeroes1': r"第.{1,2}回",
               'McNeill': r"CHAPTER \d+|Preface",
               'Fagan_Warming': r"CHAPTER \d{1,2}",
               'FortressBesieged': r"[一二三四五六七八九十〇]{1,2}",
               'PattenHKDiaries': r"(?:Sun|Mon|Tues|Wednes|Thurs|Fri|Satur)day \d{1,2}.*?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sept|Oct|Nov|Dec).*",
               'SC3': r"CHAPTER \d+",
               'RandomWalkWallStreet': r"\d{1,2}",
               'Taiping': r"Chapter \d{1,2}",
               'TwoYearsHoliday': r"CHAPTER [IVXL]+",
               'AutumnHeavenly': r"CHAPTER \d{1,2}",
               'RoweBrookQing': r"Chapter \d{1,2}",
               'Hedgehogging': r"CHAPTER [\w-]+",
               'SorkinTooBig': r"CHAPTER [A-Z-]{1,20}",
               'ChevyMissing': r"SESSION [\w-]+",
               'Contagion': r"\d",
               'Invested': r"Chapter \d{1,2}",
               'Kelly_Bitcoin': r"Chapter \d{1,2}",
               'Smith_StocksLong': r"CHAPTER.+",
               'Wuthering': r"CHAPTER [IVXL]+",
               'Crashed': r"Chapter [0-9]{1,2}",
               'FormosaBetrayed': r"#+",
               'Pantsov_CKS': r"CHAPTER \d+|\*\*\* Epilogue",
              }
    
    DregexT = {'': r"\n(",
               'HP1':   r"第.{1,3}章.*?",
               'HP2':   r"第.{1,3}章.*?",
               'HP3':   r"第.{1,3}章.*?",
               'HP4':   r"第.{1,3}章.*?",
               'HP5':   r"第.{1,3}章.*?",
               'HP6':   r"第.{1,3}章.*?",
               'HP7':   r"第.{1,3}章.*?",
               'Dubliners': r"",
               'Asimov_Foundation1': r"第.{1,3}章",
               'Asimov_Foundation2': r"第.{1,3}章　+.*?",
               'Asimov_Foundation3': r"第.{1,3}章　+.*?",
               'Asimov_FoundationPrelude': r"第.{1,3}章",
               'Asimov_FoundationEdge': r"第.{1,3}章",
               'Asimov_FoundationEarth': r"第.{1,3}章　+.*?",
               'Capitalism': r"第.{1,3}章",
               'SongIceFire1': r"第.{1,3}章.*?",
               'SongIceFire2': r"第.{1,3}章.*?",
               'SongIceFire3': r"第.{1,3}章.*?",
               'SongIceFire4': r"第.{1,3}章.*?",
               'Corton_LondonFog': r"第.{1,3}章",
               'Taipei': r"CHAPITRE \d+",
               'Patten_EastWest': r"第.{1,3}章 .+?",
               'SC3a': r"第.{1,3}章",
               'StoryStone': r"Chapter \d+",
               'Vogel_Deng': r"第.+?章",
               'Vogel_Deng_Notes': r"第.+?章 .+",
               'OrphanAsia': r"\d{1,2}\._.+?",
               'Westad_RestlessEmpire': r"第.{1,3}章：.+",
               'Tolstoy_WarPeace': r"[一二三四五六七八九零〇十白]{1,8}",
               'Yanagihara_PeopleTrees_TextPlusNotes': r"第.{1,3}章　.+",
               'DMS1': r"第.{1,3}章",
               'DMS2': r"第.{1,3}章",
               'DMS3': r"第.{1,3}章",
               'Dimbleby_LastGovernor': r"\d{1,2}",
               'Crystal': r"BOOK .{3,5}",
               'GoodEarth': r"[一二三四五六七八九〇十一──１]{1,4}",
               'Oriental': r"[一二三四五六七八九〇]、.+",
               'CondorHeroes1': r"Chapter \w+",
               'McNeill': r"第.{1,2}章|序曲|作者序",
               'Fagan_Warming': r"｜第.{1,2}章｜",
               'FortressBesieged': r"\d{1,2}",
               'PattenHKDiaries': r".{1,2}月.{1,3}日[^，。\n]*",
               'SC3': r"第.{1,3}章",
               'RandomWalkWallStreet': r"第.{1,2}章",
               'Taiping': r"第.{1,3}章",
               'TwoYearsHoliday': r"第.{1,3}章",
               'AutumnHeavenly': r"第.{1,3}章",
               'RoweBrookQing': r"第.{1,2}章",
               'Hedgehogging': r"Chapter .{1,2}",
               'SorkinTooBig': r"\d{1,2}",
               'ChevyMissing': r"第.{1,3}次心理諮商",
               'Contagion': r"\d",
               'Invested': r"第\d{1,2}章",
               'Kelly_Bitcoin': r"\d{2}",
               'Smith_StocksLong': r"第\d{1,2}章.+",
               'Wuthering': r"第.{1,3}章",
               'Crashed': r"第\d+章",
               'FormosaBetrayed': r"#+",
               'Pantsov_CKS': r"\*\*\* 第.+?章|\*\*\* 後記",
              }

    #sys.exit(0)
    
    if CONVERT_ZHS_TO_ZHT and langS == 'zhs':
        out_langS = 'zh'
        langS = 'zh'
    if CONVERT_ZHS_TO_ZHT and langT == 'zhs':
        out_langT = 'zh'
        langT = 'zh'

    for fS, fT in zip(filesS, filesT):
        
        '''
        fS: HP/HP1.en.txt, HP/HP2.en.txt, etc.
        fT: HP/HP1.zhs.txt, HP/HP2.zhs.txt, etc.
        '''
    
        #finS = f"{base}.{langS}.txt"
        #finT = f"{base}.{langT}.txt"
        finS = f"{in_folder}/{fS.name}"
        finT = f"{in_folder}/{fT.name}"
        base = fS.name.split('.')[0]
        #base = fS.name.stem

        mod = model_name.split('/')[-1]

        fon = f"{out_folder}/{base}.vecalign.n{alignment_max_size}.{mod}.{out_langS}-{out_langT}.txt"

        print(f"processing [{finS}] and [{finT}] to create [{fon}]...")
        #continue

        txtS = open(finS, "r", encoding="utf-8").read()
        if USE_REGEX_CHAPTER_SEPARATOR:
            regexS = f"\n({DregexS[base]})\n"
            chS = re.split(regexS, txtS)
        else:
            chS = [txtS]
        print(f"chS has {len(chS)} elements", flush=True)

        txtT = open(finT, "r", encoding="utf-8").read()
        if USE_REGEX_CHAPTER_SEPARATOR:
            regexT = f"\n({DregexT[base]})\n"
            chT = re.split(regexT, txtT)
        else:
            chT = [txtT]
        print(f"chT has {len(chT)} elements", flush=True)
    
        if len(chS) == len(chT):
            print("Both have the same number of elements!")
        else:
            hS = [chS[i] for i in range(len(chS)) if i % 2 == 1]
            sizeS = len(hS) 
            hT = [chT[i] for i in range(len(chT)) if i % 2 == 1]
            sizeT = len(hT)
            if sizeS > sizeT:
                for j in range(sizeS - sizeT):
                    hT.append('')
            elif sizeS < sizeT:
                for j in range(sizeT - sizeS):
                    hS.append('')
            with open(f'{out_folder}/{base}.vecalign.n{alignment_max_size}.{out_langS}-{out_langT}.ChapterMathchings.txt', 'w', encoding='utf-8') as fo:
                for s,t in zip(hS, hT):
                    fo.write(f"{s}\t{t}\n")
            sys.exit(0)

        #sys.exit(0)
    
        ch_cnt = 0
        for cS, cT in zip(chS, chT):
            
            #if cT[:2] not in ['天戰']: continue
        
            ch_cnt += 1
            print(f"processing segment [{ch_cnt}]...", flush=True)
    
            # Source    
            pS = cS.strip().split("\n")
            pS = [s.strip() for s in pS if s.strip()!='']
            sS = []
            for p in pS:
                sS.extend(sentencize(p, lang=langS))
            sS = [s.strip() for s in sS if s.strip()!='']
            ## convert source from simplified Chinese to traditional Chinese
            if CONVERT_ZHS_TO_ZHT and langS == 'zh':
                #sS = [s2tw.convert(s).replace('“','「').replace('”','」') for s in sS]
                sS = [convertChinesePunctuations(s2tw.convert(s)) for s in sS]
    
            # Target
            pT = cT.strip().split("\n")
            pT = [s.strip() for s in pT if s.strip()!='']
            sT = []
            for p in pT:
                sT.extend(sentencize(p, lang=langT))
            sT = [s.strip() for s in sT if s.strip()!='']
            ## convert target from simplified Chinese to traditional Chinese
            if CONVERT_ZHS_TO_ZHT and langT == 'zh':
                #sT = [s2tw.convert(s).replace('“','「').replace('”','」') for s in sT]
                sT = [convertChinesePunctuations(s2tw.convert(s)) for s in sT]

            with open(fon, "a", encoding="utf-8", newline="\n") as fo:
                #for score, idxE, e, idxZ, z in align(sE, sZ, alignment_max_size=alignment_max_size):
                for score, idxS, ss, idxT, tt in align(sS, sT, alignment_max_size=alignment_max_size):
                    #fo.write(f"{base}\t{score:.4f}\t{idxS}\t{ss}\t{idxT}\t{tt}\n")
                    fo.write(f"{score:.4f}\t{idxS}\t{ss}\t{idxT}\t{tt}\n")
                    fo.flush()

        print('-'*25)
        fon_xlsx = fon
        print("Creating Excel file...")
        createExcel(fon_xlsx)
        print('='*25)



#%%
